#!/usr/bin/env python3
"""
Amazon order export per marketplace — YouGuide eSIM / SIM Cards
===============================================================
Pulls EVERY order Amazon still holds for a marketplace via the SP-API **Orders
API** and writes one Excel workbook per marketplace into ``Sheets/``. Written
for the request "the list of all the orders in Sweden with ASIN numbers and
order numbers — and the same for Poland", so the default run is SE + PL:

    Sheets/orders_SE_<stamp>.xlsx      ~300 orders expected
    Sheets/orders_PL_<stamp>.xlsx      ~300 orders expected

Each workbook has two tabs:

  * **orders**   — one row per *order line*: order number, purchase date, item
                   NAME, SKU, ASIN, quantity, price, status, fulfilment.
                   An order containing two different SKUs is two rows.
  * **products** — the same data collapsed to unique NAME / SKU / ASIN with the
                   order and unit counts behind each. This is the "just name,
                   SKU and ASIN" view.

READ-ONLY. It calls getOrders / getOrderItems and touches nothing — no listing,
no sheet, no Amazon state is changed by this script.

Two things worth knowing before running:

  * **Rate limits dominate the runtime.** getOrders is 0.0167 rps (one page per
    minute at steady state, burst 20) and getOrderItems is 0.5 rps (burst 30).
    ~300 orders per market therefore means ~300 item calls at 2 s apiece — call
    it 10-12 minutes per marketplace, most of it waiting on Amazon.
  * **So everything is cached.** The order list and every order's items are
    written to ``reports_cache/orders_<CODE>.json`` as they arrive, and a
    re-run reuses them, making the second run seconds rather than minutes. The
    cache is also the crash-resume: an interrupted run picks up where it
    stopped. ``--refresh`` throws it away and re-fetches from Amazon.

Credentials, the marketplace registry and the region endpoints all come from
upload_listings.py — nothing is configured twice. Only the marketplace's own
authorization group needs a refresh token (SE and PL are both EU).

Usage:
    python fetch_orders.py                      # SE + PL, all history, both .xlsx
    python fetch_orders.py --markets SE         # one marketplace
    python fetch_orders.py --since 2026-01-01   # only orders created after a date
    python fetch_orders.py --refresh            # ignore the cache, re-fetch
    python fetch_orders.py --exclude-cancelled  # drop Cancelled orders
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sp_api.api import Orders
from sp_api.base import SellingApiException

from upload_listings import (
    LWA_CREDENTIALS,
    MARKETPLACES,
    REFRESH_TOKENS,
    REGION_ENDPOINT,
)

# The console on Windows defaults to cp1252 and dies on any non-ASCII log line.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:      # pragma: no cover - very old Python
    pass


# =============================================================================
#  CONFIGURATION
# =============================================================================

HERE = os.path.dirname(os.path.abspath(__file__))

# Where the workbooks land. The client asked for them "inside Sheets".
OUTPUT_DIR = os.path.join(HERE, "Sheets")

# Where the raw API responses are parked so a re-run costs no API calls.
CACHE_DIR = os.path.join(HERE, "reports_cache")

# Marketplaces exported when --markets is not given.
DEFAULT_MARKETS = ["SE", "PL"]

# How far back to ask for. Amazon only *keeps* a couple of years of orders in
# the Orders API, so this is deliberately earlier than the account exists: ask
# for everything and take whatever Amazon still has. If the API rejects a date
# as too old it says so, and DATE_FALLBACKS is tried in turn.
DEFAULT_SINCE = "2015-01-01"
DATE_FALLBACKS = ["2018-01-01", "2020-01-01", "2022-01-01", "2024-01-01"]

# Throttle budget (see the module docstring). Both are the documented steady
# rates with a little headroom; the burst allowance means the first calls fly.
SLEEP_ORDERS_PAGE = 2.0     # between getOrders pages      (limit 0.0167 rps)
SLEEP_ORDER_ITEMS = 2.1     # between getOrderItems calls  (limit 0.5 rps)

# Throttle / transient-error retry ladder, in seconds.
RETRY_BACKOFF = [5, 15, 45, 90, 180]

# Page size for getOrders — the API maximum, so ~300 orders is 3 pages.
MAX_RESULTS_PER_PAGE = 100

# Order statuses dropped by --exclude-cancelled.
CANCELLED_STATUSES = {"canceled", "cancelled"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("fetch_orders")


# =============================================================================
#  EXCEL COLUMNS
#  (key, header, width) — the order here is the order in the sheet. Order
#  number first because that is what the request led with, then the three
#  columns actually asked for: name, SKU, ASIN.
# =============================================================================

ORDER_COLUMNS = [
    ("order_number",   "Order number",       22),
    ("purchase_date",  "Purchase date",      20),
    ("name",           "Name",               58),
    ("sku",            "SKU",                30),
    ("asin",           "ASIN",               14),
    ("quantity",       "Qty",                6),
    ("item_price",     "Item price",         12),
    ("currency",       "Currency",           10),
    ("order_status",   "Order status",       15),
    ("fulfilment",     "Fulfilment",         12),
    ("sales_channel",  "Sales channel",      18),
    ("order_total",    "Order total",        12),
    ("last_updated",   "Last updated",       20),
]

PRODUCT_COLUMNS = [
    ("name",     "Name",   58),
    ("sku",      "SKU",    30),
    ("asin",     "ASIN",   14),
    ("orders",   "Orders", 9),
    ("units",    "Units",  9),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")


# =============================================================================
#  SP-API PLUMBING
# =============================================================================

def orders_client(code: str):
    """Orders client for a marketplace, or None when its token is unset.

    Same resolution rule as upload_listings.get_client_for_marketplace(): the
    refresh token comes from the marketplace's AUTHORIZATION GROUP and the host
    from its REGION, so SE and PL share one EU client and one EU token.
    """
    cfg = MARKETPLACES[code]
    token = REFRESH_TOKENS.get(cfg["token"], "")
    if not token:
        return None
    return Orders(
        credentials={"refresh_token": token, **LWA_CREDENTIALS},
        marketplace=REGION_ENDPOINT[cfg["region"]],
    )


def call_with_retry(fn, *args, what: str = "call", **kwargs):
    """Run an SP-API call, backing off through RETRY_BACKOFF on throttling.

    Returns the ApiResponse, or None once the ladder is exhausted — the caller
    decides whether a missing piece is fatal. Throttling is the expected error
    here, not the exception: the Orders API budget is tiny.
    """
    for attempt, wait in enumerate([0] + RETRY_BACKOFF):
        if wait:
            log.warning("  throttled/failed on %s — waiting %ss (attempt %d)",
                        what, wait, attempt)
            time.sleep(wait)
        try:
            return fn(*args, **kwargs)
        except SellingApiException as exc:
            code = getattr(exc, "code", None) or getattr(exc, "status_code", None)
            if code in (400, 403, 404):          # a real error, not a throttle
                log.error("  %s failed (%s): %s", what, code, exc)
                return None
        except Exception as exc:                 # network hiccup, JSON error…
            log.warning("  %s raised %s", what, exc)
    log.error("  %s gave up after %d attempts", what, len(RETRY_BACKOFF) + 1)
    return None


def fetch_order_list(client, code: str, since: str, log_prefix: str) -> list:
    """Every order in one marketplace, newest API page first, fully paginated.

    Tries `since` and then walks DATE_FALLBACKS if Amazon rejects the window as
    too old, so "give me everything" degrades to "everything Amazon still has"
    instead of failing.
    """
    for start in [since] + [d for d in DATE_FALLBACKS if d > since]:
        created_after = f"{start}T00:00:00Z"
        orders, token, page = [], None, 0
        while True:
            page += 1
            params = {
                "MarketplaceIds": [MARKETPLACES[code]["id"]],
                "MaxResultsPerPage": MAX_RESULTS_PER_PAGE,
            }
            if token:
                params["NextToken"] = token
            else:
                params["CreatedAfter"] = created_after

            res = call_with_retry(client.get_orders,
                                  what=f"{log_prefix} getOrders page {page}",
                                  **params)
            if res is None:
                break

            payload = res.payload or {}
            batch = payload.get("Orders", []) or []
            orders.extend(batch)
            log.info("%s page %d: %d orders (%d so far)",
                     log_prefix, page, len(batch), len(orders))

            token = payload.get("NextToken")
            if not token:
                return orders
            time.sleep(SLEEP_ORDERS_PAGE)

        if orders:                  # partial result beats nothing
            return orders
        log.warning("%s nothing returned from %s — trying a later start date",
                    log_prefix, start)
    return []


def fetch_items(client, order_id: str, log_prefix: str) -> list:
    """Every line item of one order (paginated; multi-SKU orders are rare here)."""
    items, token, guard = [], None, 0
    while guard < 20:
        guard += 1
        kwargs = {"NextToken": token} if token else {}
        res = call_with_retry(client.get_order_items, order_id,
                              what=f"{log_prefix} getOrderItems {order_id}",
                              **kwargs)
        if res is None:
            break
        payload = res.payload or {}
        items.extend(payload.get("OrderItems", []) or [])
        token = payload.get("NextToken")
        if not token:
            break
        time.sleep(SLEEP_ORDER_ITEMS)
    return items


# =============================================================================
#  CACHE  —  the API budget is the scarce resource, so nothing is fetched twice
# =============================================================================

def cache_path(code: str) -> str:
    return os.path.join(CACHE_DIR, f"orders_{code}.json")


EMPTY_CACHE = {"orders": [], "items": {}, "since": ""}


def load_cache(code: str, refresh: bool, since: str) -> dict:
    """{"orders": [...], "items": {order_id: [...]}, "since": …} for a market.

    The cached ORDER LIST only covers the window it was fetched with, so a run
    asking for MORE history than the cache holds must re-fetch the list — else
    a narrow canary run would quietly cap the full export. The cached ITEMS are
    keyed by order id and stay valid whatever the window, so they are kept.
    """
    if refresh:
        return dict(EMPTY_CACHE, items={})
    path = cache_path(code)
    if not os.path.exists(path):
        return dict(EMPTY_CACHE, items={})
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        cache = {"orders": data.get("orders", []),
                 "items":  data.get("items", {}),
                 "since":  data.get("since", "")}
    except (OSError, ValueError) as exc:
        log.warning("cache %s unreadable (%s) — ignoring it", path, exc)
        return dict(EMPTY_CACHE, items={})

    if cache["orders"] and cache["since"] > since:
        log.info("cached order list starts at %s, this run wants %s — re-fetching "
                 "the list (item detail is kept)", cache["since"], since)
        cache["orders"] = []
    return cache


def save_cache(code: str, cache: dict) -> None:
    os.makedirs(CACHE_DIR, exist_ok=True)
    tmp = cache_path(code) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump({"fetched": datetime.now(timezone.utc).isoformat(),
                   "since": cache.get("since", ""),
                   "orders": cache["orders"], "items": cache["items"]},
                  fh, ensure_ascii=False)
    os.replace(tmp, cache_path(code))


# =============================================================================
#  SHAPING
# =============================================================================

def parse_dt(value: str):
    """Amazon's ISO timestamp → a naive UTC datetime Excel can format."""
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return value


def money(block: dict):
    """{'CurrencyCode': 'SEK', 'Amount': '129.00'} → (129.0, 'SEK')."""
    if not isinstance(block, dict):
        return "", ""
    try:
        return float(block.get("Amount", "") or 0), block.get("CurrencyCode", "")
    except (TypeError, ValueError):
        return "", block.get("CurrencyCode", "")


def build_rows(orders: list, items_by_order: dict, exclude_cancelled: bool) -> list:
    """One row per order LINE. Orders with no items still get a row.

    A Pending order returns items without pricing (Amazon withholds it until
    payment authorises), and an order Amazon has purged items for returns none
    at all — both still carry an order number the client asked for, so they are
    kept with the missing fields blank rather than silently dropped.
    """
    rows = []
    for order in orders:
        status = order.get("OrderStatus", "")
        if exclude_cancelled and status.lower() in CANCELLED_STATUSES:
            continue

        order_id  = order.get("AmazonOrderId", "")
        total, cur = money(order.get("OrderTotal"))
        base = {
            "order_number":  order_id,
            "purchase_date": parse_dt(order.get("PurchaseDate", "")),
            "order_status":  status,
            "fulfilment":    order.get("FulfillmentChannel", ""),
            "sales_channel": order.get("SalesChannel", ""),
            "order_total":   total,
            "last_updated":  parse_dt(order.get("LastUpdateDate", "")),
        }

        lines = items_by_order.get(order_id) or []
        if not lines:
            rows.append({**base, "name": "", "sku": "", "asin": "",
                         "quantity": "", "item_price": "", "currency": cur})
            continue

        for item in lines:
            price, item_cur = money(item.get("ItemPrice"))
            rows.append({
                **base,
                "name":       item.get("Title", ""),
                "sku":        item.get("SellerSKU", ""),
                "asin":       item.get("ASIN", ""),
                "quantity":   item.get("QuantityOrdered", ""),
                "item_price": price,
                "currency":   item_cur or cur,
            })

    rows.sort(key=lambda r: (str(r["purchase_date"]), r["order_number"]))
    return rows


def build_products(rows: list) -> list:
    """Collapse the order lines to unique NAME / SKU / ASIN with their counts."""
    seen: dict = {}
    for row in rows:
        if not (row["sku"] or row["asin"]):
            continue
        key = (row["sku"], row["asin"])
        entry = seen.setdefault(key, {"name": row["name"], "sku": row["sku"],
                                      "asin": row["asin"], "orders": set(),
                                      "units": 0})
        entry["orders"].add(row["order_number"])
        try:
            entry["units"] += int(row["quantity"] or 0)
        except (TypeError, ValueError):
            pass
        if not entry["name"]:
            entry["name"] = row["name"]

    products = [{**e, "orders": len(e["orders"])} for e in seen.values()]
    products.sort(key=lambda p: (-p["units"], p["name"]))
    return products


# =============================================================================
#  EXCEL
# =============================================================================

def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def write_excel(code: str, rows: list, products: list, path: str,
                since: str) -> None:
    """Workbook with the per-line 'orders' tab and the 'products' summary tab."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "orders"
    ws.append([label for _, label, _ in ORDER_COLUMNS])
    col_of = {key: i + 1 for i, (key, _, _) in enumerate(ORDER_COLUMNS)}
    for key, _, width in ORDER_COLUMNS:
        ws.column_dimensions[get_column_letter(col_of[key])].width = width

    for row in rows:
        ws.append([row.get(key, "") for key, _, _ in ORDER_COLUMNS])
        r = ws.max_row
        for key in ("purchase_date", "last_updated"):
            ws.cell(row=r, column=col_of[key]).number_format = "yyyy-mm-dd hh:mm"
        for key in ("item_price", "order_total"):
            ws.cell(row=r, column=col_of[key]).number_format = "0.00"
        # Grey the line out when Amazon returned no item detail for the order.
        if not row.get("asin") and not row.get("sku"):
            for col in range(1, len(ORDER_COLUMNS) + 1):
                ws.cell(row=r, column=col).fill = GREY_FILL
    style_header(ws, len(ORDER_COLUMNS))

    ps = wb.create_sheet("products")
    ps.append([label for _, label, _ in PRODUCT_COLUMNS])
    for i, (_, _, width) in enumerate(PRODUCT_COLUMNS):
        ps.column_dimensions[get_column_letter(i + 1)].width = width
    for product in products:
        ps.append([product.get(key, "") for key, _, _ in PRODUCT_COLUMNS])
    style_header(ps, len(PRODUCT_COLUMNS))

    ps.append([])
    ps.append(["Marketplace", f"{code} (amazon {MARKETPLACES[code]['id']})"])
    ps.append(["Orders", len({r['order_number'] for r in rows})])
    ps.append(["Order lines", len(rows)])
    ps.append(["Distinct products", len(products)])
    ps.append(["Created after", since])
    ps.append(["Generated (UTC)", datetime.now(timezone.utc).replace(tzinfo=None)])
    ps.cell(row=ps.max_row, column=2).number_format = "yyyy-mm-dd hh:mm"

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# =============================================================================
#  PER-MARKETPLACE RUN
# =============================================================================

def export_market(code: str, args, stamp: str) -> str:
    """Fetch (or reuse) one marketplace's orders and write its workbook."""
    prefix = f"[{code}]"
    log.info("%s ── %s ──", prefix, MARKETPLACES[code]["id"])

    cache = load_cache(code, args.refresh, args.since)
    client = orders_client(code)

    if client is None:
        log.error("%s no refresh token for authorization group %s — skipped",
                  prefix, MARKETPLACES[code]["token"])
        if not cache["orders"]:
            return ""
        log.warning("%s writing from cache only", prefix)

    if client is not None and not cache["orders"]:
        cache["orders"] = fetch_order_list(client, code, args.since, prefix)
        cache["since"] = args.since
        save_cache(code, cache)
    elif cache["orders"]:
        log.info("%s %d orders from cache (--refresh to re-fetch)",
                 prefix, len(cache["orders"]))

    orders = cache["orders"]
    if not orders:
        log.error("%s no orders returned", prefix)
        return ""

    # Item detail, one call per order — the slow part, hence the running save.
    missing = [o["AmazonOrderId"] for o in orders
               if o.get("AmazonOrderId") and o["AmazonOrderId"] not in cache["items"]]
    if missing and client is not None:
        log.info("%s fetching items for %d orders (~%d min)",
                 prefix, len(missing), max(1, round(len(missing) * SLEEP_ORDER_ITEMS / 60)))
        for i, order_id in enumerate(missing, 1):
            cache["items"][order_id] = fetch_items(client, order_id, prefix)
            if i % 25 == 0 or i == len(missing):
                save_cache(code, cache)
                log.info("%s items %d/%d", prefix, i, len(missing))
            if i < len(missing):
                time.sleep(SLEEP_ORDER_ITEMS)
        save_cache(code, cache)
    elif missing:
        log.warning("%s %d orders have no cached items and no client to fetch them",
                    prefix, len(missing))

    rows = build_rows(orders, cache["items"], args.exclude_cancelled)
    products = build_products(rows)

    path = os.path.join(OUTPUT_DIR, f"orders_{code}_{stamp}.xlsx")
    write_excel(code, rows, products, path, args.since)
    log.info("%s %d orders, %d lines, %d products → %s",
             prefix, len({r['order_number'] for r in rows}), len(rows),
             len(products), path)
    return path


# =============================================================================
#  ENTRY POINT
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export every Amazon order for a marketplace to Excel "
                    "(order number, name, SKU, ASIN). Read-only.")
    parser.add_argument("--markets", default=",".join(DEFAULT_MARKETS),
                        help="comma list of marketplace codes (default SE,PL)")
    parser.add_argument("--since", default=DEFAULT_SINCE,
                        help="only orders created after this date, YYYY-MM-DD "
                             f"(default {DEFAULT_SINCE} = everything Amazon has)")
    parser.add_argument("--refresh", action="store_true",
                        help="ignore reports_cache and re-fetch from Amazon")
    parser.add_argument("--exclude-cancelled", action="store_true",
                        help="drop Cancelled orders (they are included by default)")
    args = parser.parse_args()

    codes = [c.strip().upper() for c in args.markets.split(",") if c.strip()]
    unknown = [c for c in codes if c not in MARKETPLACES]
    if unknown:
        parser.error(f"unknown marketplace code(s): {', '.join(unknown)}")

    try:
        datetime.strptime(args.since, "%Y-%m-%d")
    except ValueError:
        parser.error("--since must be YYYY-MM-DD")

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    written = []
    for code in codes:
        try:
            path = export_market(code, args, stamp)
        except KeyboardInterrupt:
            log.warning("interrupted — the cache holds what was fetched so far; "
                        "re-run to resume")
            raise
        if path:
            written.append(path)

    log.info("done — %d workbook(s) written", len(written))
    for path in written:
        log.info("  %s", path)


if __name__ == "__main__":
    main()
