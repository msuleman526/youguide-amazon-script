#!/usr/bin/env python3
"""
Catalogue enforcement — the master sheet is the ONLY thing allowed on Amazon
============================================================================
Audits EVERY marketplace against the master product list and removes every
listing that is not on it.

WHY THIS EXISTS
---------------
The account accumulated ~1,950 listings while the real catalogue is 453 SKUs.
The surplus is almost entirely OLD-NAMING-CONVENTION DUPLICATES: `HK_5_30` and
`HONG-KONG-5GB-30D` are the *same product on the same ASIN* (B0FR9GXYYL), so
Amazon treats them as competing offers on one detail page and shows the cheapest
one. The stale £1.99 offer won the buy box, a customer bought a 5GB eSIM for
£1.99, and the price gap tripped "potential high pricing error" suppressions
that took the correct listings down with it.

Deleting the surplus is therefore not housekeeping — it is the fix.

WHAT IT DOES
------------
For every marketplace (or the ones named with --markets):

  1. Pulls the FULL merchant listings report (GET_MERCHANT_LISTINGS_ALL_DATA) —
     the same source fetch_listings.py uses, because searchListingsItems will
     not paginate for this account.
  2. Classifies every live SKU against the allow-list:
       KEEP          — SKU is in the master xlsx
       KEEP-PARENT   — SKU is a variation parent referenced by the Google Sheet
       KEEP-PATTERN  — SKU matched KEEP_REGEX (parent-shaped, e.g. *-ESIM)
       DELETE        — everything else
  3. Reports the GAPS too: master SKUs that are missing from that marketplace.
     This is the other half of the problem — FLEX cannot build an FBA shipment
     because SKUs like UNITED-KINGDOM-5GB-30D are "MSKU not found" on amazon.pl.
  4. Writes two CSVs: the deletion plan and a SKU x marketplace coverage matrix.
  5. With --live, deletes the DELETE set via SP-API.

SAFETY
------
  * Previews by default. Nothing is written to Amazon without --live.
  * A typed "DELETE" confirmation is required for large live runs (--yes skips).
  * Deletions are journalled to logs/purge_ledger.csv and skipped on re-run, so
    an interrupted run resumes instead of re-issuing thousands of calls.
  * If the report's SKU column cannot be identified the run ABORTS — it will
    never guess which column holds the SKU and delete the wrong thing.
  * --limit N converts any run into a canary: N deletions per marketplace.

USAGE
-----
    python purge_listings.py                        # preview everything, no writes
    python purge_listings.py --markets UK           # one marketplace
    python purge_listings.py --report-only          # audit + CSVs, skip the plan summary
    python purge_listings.py --markets UK --live --limit 5    # canary
    python purge_listings.py --live                 # the real cleanup
    python purge_listings.py --refresh              # ignore today's cached reports

Credentials, marketplace registry and clients all come from upload_listings, so
there is still exactly one place where secrets and IDs live.
"""

from __future__ import annotations

import argparse
import csv
import io
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from sp_api.api import Reports
from sp_api.base import SellingApiException
from sp_api.base.reportTypes import ReportType

from upload_listings import (
    LWA_CREDENTIALS,
    MARKETPLACES,
    REFRESH_TOKENS,
    REGION_ENDPOINT,
    SELLER_IDS,
    get_client_for_marketplace,
)

# The console on Windows defaults to cp1252 and dies on any non-ASCII log line.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:      # pragma: no cover - very old Python
    pass


# =============================================================================
#  CONFIGURATION
# =============================================================================

HERE = os.path.dirname(os.path.abspath(__file__))

# The client's definitive product list. Every SKU in it is allowed on Amazon;
# nothing else is. Override with --master.
MASTER_FILE = os.path.join(
    HERE, "Sheets", "YOUGUIDE poducts FOR every MARKET place 1ST DISTRIBUTION (1).xlsx"
)

# Header text that identifies the SKU column in the master workbook. The first
# worksheet containing one of these headers is used.
MASTER_SKU_HEADERS = ("sku", "seller-sku", "seller sku", "msku")

# Variation PARENTS are not in the master list (it only lists sellable children)
# but deleting one destroys the variation family, so they are protected. The
# authoritative parent set is the Google Sheet's column B, and it is COMPLETE —
# one parent per place, for every place in the master list.
#
# This regex is therefore only a fallback for --no-sheet runs. It is deliberately
# NOT applied when the sheet was read: an orphan parent whose children are all
# gone from the catalogue (e.g. HAJJ-LUXURY-ESIM, an "Incomplete" parent for a
# product that no longer exists) is exactly the kind of leftover to remove, and
# the pattern would otherwise keep it alive forever.
KEEP_REGEX = re.compile(r"-ESIM$", re.IGNORECASE)

# Reports are large and slow to generate; today's copy is reused unless
# --refresh is passed.
REPORT_CACHE_DIR = os.path.join(HERE, "reports_cache")
REPORT_TYPE = ReportType.GET_MERCHANT_LISTINGS_ALL_DATA

# Real FBA stock per SKU. Deleting a listing that still holds units at a
# fulfilment centre strands that inventory, and the merchant listings report
# cannot tell you it is there (see fba_quantities).
FBA_REPORT_TYPE = ReportType.GET_FBA_MYI_UNSUPPRESSED_INVENTORY_DATA

REPORT_POLL_SECONDS = 5
REPORT_TIMEOUT_SECONDS = 900        # big accounts take a while
HTTP_RETRIES = 5
HTTP_BACKOFF_S = 3

# Listings Items deleteListingsItem is rate limited at 5 requests/second.
DELETE_SLEEP_SECONDS = 0.25
DELETE_RETRIES = 4
DELETE_BACKOFF_S = 5

# Live runs above this many deletions demand a typed confirmation.
CONFIRM_THRESHOLD = 25

OUTPUT_DIR = os.path.join(HERE, "logs")
LEDGER_FILE = os.path.join(OUTPUT_DIR, "purge_ledger.csv")

# Report column aliases. Amazon localises some report headers, so each logical
# field lists every spelling seen. A missing SKU column aborts the run.
COLUMN_ALIASES = {
    "sku":      ("seller-sku", "sku", "msku", "seller sku"),
    "asin":     ("asin1", "asin"),
    "name":     ("item-name", "item name"),
    "price":    ("price",),
    "quantity": ("quantity",),
    "status":   ("status",),
    "channel":  ("fulfillment-channel", "fulfilment-channel"),
    "opened":   ("open-date", "open date"),
}


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("purge")


# =============================================================================
#  HELPERS
# =============================================================================

def now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def norm_sku(value: object) -> str:
    """Comparison key for a SKU: trimmed, upper-cased, inner whitespace dropped.

    Amazon SKUs are case-sensitive on the wire, so the ORIGINAL spelling is what
    gets sent to the API — this key is only ever used for set membership, which
    keeps a stray lower-case row in the master file from causing a deletion.
    """
    return re.sub(r"\s+", "", str(value or "")).upper()


def with_retry(fn, *args, **kwargs):
    """Call fn, retrying transient network failures (flaky TLS handshakes)."""
    for attempt in range(1, HTTP_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except SellingApiException:
            raise                       # real API errors are the caller's problem
        except Exception as exc:
            if attempt == HTTP_RETRIES:
                raise
            log.warning("Network hiccup (%s) - retry %d/%d",
                        exc.__class__.__name__, attempt, HTTP_RETRIES)
            time.sleep(HTTP_BACKOFF_S)


# =============================================================================
#  THE ALLOW-LIST
# =============================================================================

def load_master(path: str) -> dict:
    """Read the master workbook into {norm_sku: {sku, gtin, description}}.

    Finds the first worksheet with a recognisable SKU header rather than
    hard-coding a tab name, so a re-exported file with a renamed tab still works.
    """
    if not os.path.exists(path):
        raise SystemExit(f"Master product list not found: {path}\n"
                         f"Pass the right file with --master.")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        header = [str(c or "").strip().lower()
                  for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        sku_col = next((i for i, h in enumerate(header) if h in MASTER_SKU_HEADERS), None)
        if sku_col is None:
            continue

        def find(*names):
            return next((i for i, h in enumerate(header)
                         if any(n in h for n in names)), None)

        gtin_col = find("gtin", "ean", "barcode", "product code")
        desc_col = find("description", "product description", "title")

        master: dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or sku_col >= len(row):
                continue
            raw = str(row[sku_col] or "").strip()
            if not raw:
                continue
            key = norm_sku(raw)
            if key in master:
                log.warning("Master list has SKU %s more than once - keeping the first", raw)
                continue
            master[key] = {
                "sku":  raw,
                "gtin": str(row[gtin_col]).strip() if gtin_col is not None
                        and gtin_col < len(row) and row[gtin_col] is not None else "",
                "description": str(row[desc_col]).strip() if desc_col is not None
                        and desc_col < len(row) and row[desc_col] is not None else "",
            }

        log.info("Master list: %d SKUs from %r (tab %r)",
                 len(master), os.path.basename(path), ws.title)
        return master

    raise SystemExit(
        f"No worksheet in {path} has a SKU column "
        f"(looked for a header in {MASTER_SKU_HEADERS})."
    )


def load_sheet_parents(master: dict) -> set:
    """Variation parent SKUs from the Google Sheet, for rows that are in master.

    A parent listing is not sellable and never appears in the master product
    list, but deleting one dissolves the variation family (and the child pages
    lose their parent ASIN). Only parents belonging to a KEPT child are
    protected, so an obsolete family is still cleaned up.
    """
    from upload_listings import COL_PARENT_SKU, COL_SKU, get_worksheet

    ws = get_worksheet()
    rows = ws.get_all_values()[1:]          # row 1 is the header
    parents = set()
    for row in rows:
        if len(row) <= COL_PARENT_SKU:
            continue
        child = norm_sku(row[COL_SKU] if len(row) > COL_SKU else "")
        parent = str(row[COL_PARENT_SKU] or "").strip()
        if parent and child in master:
            parents.add(norm_sku(parent))
    log.info("Google Sheet: %d variation parent SKUs protected", len(parents))
    return parents


# =============================================================================
#  REPORTS API — what is actually live, per marketplace
# =============================================================================

def reports_client(code: str) -> Optional[Reports]:
    """A Reports client bound to the marketplace's region endpoint and token."""
    cfg = MARKETPLACES[code]
    token = REFRESH_TOKENS.get(cfg["token"], "")
    if not token:
        return None
    return Reports(
        credentials={"refresh_token": token, **LWA_CREDENTIALS},
        marketplace=REGION_ENDPOINT[cfg["region"]],
    )


def cache_path(code: str, kind: str) -> str:
    day = datetime.now(timezone.utc).strftime("%Y%m%d")
    return os.path.join(REPORT_CACHE_DIR, f"{kind}_{code}_{day}.tsv")


def fetch_report(code: str, refresh: bool = False,
                 report_type=REPORT_TYPE, kind: str = "merchant_listings",
                 quiet: bool = False) -> Optional[str]:
    """Return a report's TSV for one marketplace, or None on failure.

    Cached per marketplace per day: a full audit of 15 marketplaces costs 15
    report generations, and re-running the plan after a partial delete should
    not pay that twice.
    """
    os.makedirs(REPORT_CACHE_DIR, exist_ok=True)
    path = cache_path(code, kind)
    if os.path.exists(path) and not refresh:
        log.info("[%s] using cached %s report", code, kind)
        return open(path, encoding="utf-8", errors="replace").read()

    client = reports_client(code)
    if client is None:
        log.warning("[%s] no refresh token for auth group %r - skipped",
                    code, MARKETPLACES[code]["token"])
        return None

    mp_id = MARKETPLACES[code]["id"]
    log.info("[%s] requesting %s report...", code, kind)
    try:
        resp = with_retry(
            client.create_report,
            reportType=report_type if isinstance(report_type, str) else report_type.value,
            marketplaceIds=[mp_id],
        )
        report_id = (resp.payload or {}).get("reportId")
        if not report_id:
            log.error("[%s] create_report returned no reportId: %s", code, resp.payload)
            return None

        waited = 0
        while True:
            payload = (with_retry(client.get_report, report_id).payload or {})
            status = payload.get("processingStatus", "UNKNOWN")
            if status == "DONE":
                doc_id = payload.get("reportDocumentId")
                break
            if status in ("CANCELLED", "FATAL"):
                log.error("[%s] report %s ended %s", code, report_id, status)
                return None
            if waited >= REPORT_TIMEOUT_SECONDS:
                log.error("[%s] report %s still %s after %ds - giving up",
                          code, report_id, status, waited)
                return None
            time.sleep(REPORT_POLL_SECONDS)
            waited += REPORT_POLL_SECONDS
            if waited % 30 == 0:
                log.info("[%s] report %s: %s (%ds)", code, report_id, status, waited)

        text = (with_retry(client.get_report_document, doc_id, download=True)
                .payload or {}).get("document")
        if text is None:
            log.error("[%s] report document %s had no content", code, doc_id)
            return None

    except SellingApiException as exc:
        level = log.warning if quiet else log.error
        level("[%s] SP-API error fetching %s report: %s", code, kind, exc)
        return None

    with open(path, "w", encoding="utf-8", errors="replace", newline="") as fh:
        fh.write(text)
    log.info("[%s] %s report cached to %s", code, kind, os.path.basename(path))
    return text


def fba_quantities(code: str, refresh: bool = False) -> Optional[dict]:
    """{norm_sku: fulfillable units} from the FBA inventory report, or None.

    The merchant listings report's `quantity` column is MERCHANT-fulfilled stock,
    so it reads 0 for every FBA listing — which would make a purge look risk-free
    while it is in fact about to strand real units at a fulfilment centre (the
    screenshot of HK_5_30 showed 8 available against a reported quantity of 0).
    This second report is the only place the real number lives.

    Returns None when the report is unavailable (some marketplaces do not support
    it) so the caller can say "unknown" rather than "zero".
    """
    text = fetch_report(code, refresh=refresh, report_type=FBA_REPORT_TYPE,
                        kind="fba_inventory", quiet=True)
    if text is None:
        log.warning("[%s] FBA inventory report unavailable - stock levels unknown", code)
        return None

    reader = csv.reader(io.StringIO(text), delimiter="\t")
    rows = [r for r in reader if r]
    if not rows:
        return None
    header = [h.strip().lower() for h in rows[0]]
    sku_i = next((header.index(a) for a in ("sku", "seller-sku", "msku") if a in header), None)
    qty_i = next((header.index(a) for a in
                  ("afn-fulfillable-quantity", "afn-warehouse-quantity",
                   "afn-total-quantity", "quantity-available") if a in header), None)
    if sku_i is None or qty_i is None:
        log.warning("[%s] FBA report header not recognised (%s) - stock unknown",
                    code, ", ".join(header[:12]))
        return None

    quantities: dict = {}
    for row in rows[1:]:
        if sku_i >= len(row):
            continue
        try:
            qty = int(float(row[qty_i])) if qty_i < len(row) and row[qty_i] else 0
        except ValueError:
            qty = 0
        key = norm_sku(row[sku_i])
        if key:
            quantities[key] = max(qty, quantities.get(key, 0))
    return quantities


def parse_report(code: str, text: str) -> list:
    """TSV -> list of dicts with the fields this script needs.

    Aborts the marketplace (returns []) if the SKU column is unrecognisable,
    rather than falling back to a positional guess — a wrong guess here deletes
    the wrong listings.
    """
    reader = csv.reader(io.StringIO(text), delimiter="\t")
    rows = [r for r in reader if r]
    if not rows:
        log.warning("[%s] report is empty", code)
        return []

    header = [h.strip().lower() for h in rows[0]]
    index = {}
    for field, aliases in COLUMN_ALIASES.items():
        index[field] = next((header.index(a) for a in aliases if a in header), None)

    if index["sku"] is None:
        log.error("[%s] ABORTED - no SKU column in the report. Header was: %s",
                  code, ", ".join(header[:40]))
        return []

    def cell(row, field):
        i = index[field]
        if i is None or i >= len(row):
            return ""
        return str(row[i] or "").strip()

    listings = []
    for row in rows[1:]:
        sku = cell(row, "sku")
        if not sku:
            continue
        listings.append({
            "sku":      sku,
            "key":      norm_sku(sku),
            "asin":     cell(row, "asin"),
            "name":     cell(row, "name"),
            "price":    cell(row, "price"),
            "quantity": cell(row, "quantity"),
            "status":   cell(row, "status"),
            "channel":  cell(row, "channel"),
            "opened":   cell(row, "opened"),
        })
    return listings


# =============================================================================
#  CLASSIFICATION
# =============================================================================

def classify(listing: dict, master: dict, parents: set,
             use_pattern: bool) -> tuple[str, str]:
    """(decision, reason) for one live listing.

    use_pattern is only True when the Google Sheet was NOT read — see KEEP_REGEX.
    """
    key = listing["key"]
    if key in master:
        return "KEEP", "in master list"
    if key in parents:
        return "KEEP-PARENT", "variation parent in Google Sheet"
    if use_pattern and KEEP_REGEX.search(listing["sku"]):
        return "KEEP-PATTERN", f"matched KEEP_REGEX {KEEP_REGEX.pattern} (no sheet)"
    return "DELETE", "not in master list"


def audit_market(code: str, master: dict, parents: set, refresh: bool,
                 use_pattern: bool) -> Optional[dict]:
    """Full picture for one marketplace: what to keep, delete, and what's missing."""
    text = fetch_report(code, refresh=refresh)
    if text is None:
        return None
    listings = parse_report(code, text)
    if not listings:
        return None

    # Real FBA stock, which the merchant listings report does NOT carry.
    fba = fba_quantities(code, refresh=refresh)

    keep, delete = [], []
    for item in listings:
        decision, reason = classify(item, master, parents, use_pattern)
        item["decision"], item["reason"] = decision, reason
        item["fba_qty"] = fba.get(item["key"], "" if fba is None else 0)
        (delete if decision == "DELETE" else keep).append(item)

    present = {i["key"] for i in listings}
    missing = [master[k]["sku"] for k in master if k not in present]

    stocked = [d for d in delete if isinstance(d["fba_qty"], int) and d["fba_qty"] > 0]

    log.info("[%s] %d live | keep %d | DELETE %d (%s) | missing %d of %d",
             code, len(listings), len(keep), len(delete),
             "FBA stock unknown" if fba is None
             else f"{len(stocked)} holding {sum(d['fba_qty'] for d in stocked)} FBA units",
             len(missing), len(master))
    return {
        "code": code, "listings": listings, "keep": keep, "delete": delete,
        "missing": sorted(missing), "stocked": stocked, "fba_known": fba is not None,
    }


# =============================================================================
#  OUTPUT
# =============================================================================

def write_plan_csv(results: list, stamp: str) -> str:
    path = os.path.join(OUTPUT_DIR, f"cleanup_plan_{stamp}.csv")
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["marketplace", "decision", "reason", "sku", "asin", "item-name",
                    "price", "fba-units", "status", "fulfillment-channel", "open-date"])
        for res in results:
            for item in sorted(res["listings"],
                               key=lambda i: (i["decision"] != "DELETE", i["sku"])):
                w.writerow([res["code"], item["decision"], item["reason"], item["sku"],
                            item["asin"], item["name"], item["price"],
                            item["fba_qty"] if item["fba_qty"] != "" else "unknown",
                            item["status"], item["channel"], item["opened"]])
    return path


def write_coverage_csv(results: list, master: dict, stamp: str) -> str:
    """SKU x marketplace matrix - the direct answer to FLEX's 'MSKU not found'.

    Each cell is the listing's Amazon status, or MISSING where the master SKU has
    no listing in that marketplace at all.
    """
    path = os.path.join(OUTPUT_DIR, f"coverage_{stamp}.csv")
    codes = [r["code"] for r in results]
    by_market = {r["code"]: {i["key"]: i for i in r["listings"]} for r in results}

    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sku", "description", "markets_missing"] + codes)
        for key in sorted(master, key=lambda k: master[k]["sku"]):
            cells, missing_count = [], 0
            for code in codes:
                item = by_market[code].get(key)
                if item is None:
                    cells.append("MISSING")
                    missing_count += 1
                else:
                    cells.append(item["status"] or "PRESENT")
            w.writerow([master[key]["sku"], master[key]["description"],
                        missing_count] + cells)
    return path


def print_summary(results: list, master: dict) -> None:
    log.info("%s", "=" * 82)
    log.info("%-4s %8s %8s %8s %9s %9s %8s", "MKT", "LIVE", "KEEP", "DELETE",
             "ACTIVE", "FBA UNITS", "MISSING")
    log.info("%s", "-" * 82)
    for res in results:
        active = sum(1 for d in res["delete"] if d["status"].lower().startswith("active"))
        units = sum(d["fba_qty"] for d in res["stocked"])
        log.info("%-4s %8d %8d %8d %9d %9s %8d", res["code"], len(res["listings"]),
                 len(res["keep"]), len(res["delete"]), active,
                 units if res["fba_known"] else "?", len(res["missing"]))
    log.info("%s", "-" * 82)
    log.info("%-4s %8d %8d %8d %9d %9d", "ALL",
             sum(len(r["listings"]) for r in results),
             sum(len(r["keep"]) for r in results),
             sum(len(r["delete"]) for r in results),
             sum(1 for r in results for d in r["delete"]
                 if d["status"].lower().startswith("active")),
             sum(d["fba_qty"] for r in results for d in r["stocked"]))
    log.info("Master list holds %d SKUs; every marketplace should end up with exactly "
             "that many KEEP rows plus its variation parents.", len(master))
    log.info("ACTIVE = live offers being deleted; these are the ones competing with a "
             "real SKU on the same ASIN. FBA UNITS = stock that will be stranded.")
    log.info("%s", "=" * 82)


def print_samples(results: list, limit: int = 12) -> None:
    """Show a slice of what would be deleted - the last check before --live."""
    for res in results:
        if not res["delete"]:
            continue
        # Active offers first: they are what is actually costing money right now.
        ordered = sorted(res["delete"],
                         key=lambda i: (not i["status"].lower().startswith("active"),
                                        i["sku"]))
        log.info("[%s] would DELETE %d, first %d (ACTIVE listed first):",
                 res["code"], len(res["delete"]), min(limit, len(res["delete"])))
        for item in ordered[:limit]:
            log.info("      %-28s %-12s %-9s fba=%-5s %-8s %s", item["sku"],
                     item["asin"], item["status"],
                     item["fba_qty"] if item["fba_qty"] != "" else "?",
                     item["price"] or "-", item["name"][:48])
        if len(res["delete"]) > limit:
            log.info("      ... and %d more (see the CSV)", len(res["delete"]) - limit)


# =============================================================================
#  DELETION
# =============================================================================

def load_ledger() -> set:
    """(marketplace, norm_sku) pairs already deleted successfully in a past run."""
    done = set()
    if not os.path.exists(LEDGER_FILE):
        return done
    with open(LEDGER_FILE, encoding="utf-8", newline="") as fh:
        for row in csv.reader(fh):
            if len(row) >= 4 and row[3] == "OK":
                done.add((row[1], norm_sku(row[2])))
    if done:
        log.info("Ledger: %d SKUs already deleted in an earlier run - they are skipped",
                 len(done))
    return done


def append_ledger(code: str, sku: str, result: str, detail: str) -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    new = not os.path.exists(LEDGER_FILE)
    with open(LEDGER_FILE, "a", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        if new:
            w.writerow(["timestamp", "marketplace", "sku", "result", "detail"])
        w.writerow([datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    code, sku, result, detail])


def delete_sku(code: str, sku: str) -> tuple[bool, str]:
    """Delete one SKU from one marketplace. Returns (ok, detail).

    Note this removes the OFFER, not the ASIN: the detail page survives, the
    listing stops being one of our offers. That is exactly what is wanted for
    the old-naming duplicates that are undercutting the real SKUs.
    """
    client = get_client_for_marketplace(code)
    if client is None:
        return False, "no client (missing token)"
    cfg = MARKETPLACES[code]
    seller_id = SELLER_IDS.get(cfg["token"], "")
    if not seller_id:
        return False, f"no seller id for auth group {cfg['token']}"

    for attempt in range(1, DELETE_RETRIES + 1):
        try:
            payload = (client.delete_listings_item(
                sellerId=seller_id, sku=sku, marketplaceIds=[cfg["id"]],
            ).payload or {})
            status = payload.get("status", "UNKNOWN")
            if status in ("ACCEPTED", "VALID"):
                return True, status
            issues = "; ".join(i.get("message", str(i))
                               for i in payload.get("issues", [])[:5])
            return False, f"{status} | {issues}"
        except SellingApiException as exc:
            text = str(exc)
            # 429s and 5xx are worth another go; anything else is a real refusal.
            if attempt < DELETE_RETRIES and ("429" in text or "throttl" in text.lower()
                                             or "50" in text[:6]):
                time.sleep(DELETE_BACKOFF_S * attempt)
                continue
            return False, text[:200]
        except Exception as exc:                        # transient network
            if attempt < DELETE_RETRIES:
                time.sleep(DELETE_BACKOFF_S * attempt)
                continue
            return False, f"{exc.__class__.__name__}: {exc}"[:200]
    return False, "retries exhausted"


def run_deletions(results: list, limit: Optional[int]) -> None:
    """Delete every DELETE row, ledgered and throttled, with a retry sweep.

    Failures are retried once at the very end: a variation PARENT cannot be
    deleted while it still has children, so a parent that failed on the first
    pass usually succeeds after its children are gone.
    """
    done = load_ledger()
    failures: list = []
    deleted = skipped = 0

    for res in results:
        code = res["code"]
        targets = res["delete"][:limit] if limit else res["delete"]
        if not targets:
            continue
        log.info("[%s] deleting %d listing(s)%s...", code, len(targets),
                 f" (capped by --limit {limit})" if limit else "")
        for n, item in enumerate(targets, 1):
            if (code, item["key"]) in done:
                skipped += 1
                continue
            ok, detail = delete_sku(code, item["sku"])
            append_ledger(code, item["sku"], "OK" if ok else "FAIL", detail)
            if ok:
                deleted += 1
                done.add((code, item["key"]))
            else:
                failures.append((code, item))
                log.warning("[%s] FAILED %s - %s", code, item["sku"], detail)
            if n % 50 == 0:
                log.info("[%s] %d/%d processed", code, n, len(targets))
            time.sleep(DELETE_SLEEP_SECONDS)

    if failures:
        log.info("Retry sweep for %d failure(s) (parents usually clear once their "
                 "children are gone)...", len(failures))
        still_failing = []
        for code, item in failures:
            ok, detail = delete_sku(code, item["sku"])
            append_ledger(code, item["sku"], "OK" if ok else "FAIL", f"retry: {detail}")
            if ok:
                deleted += 1
            else:
                still_failing.append((code, item["sku"], detail))
            time.sleep(DELETE_SLEEP_SECONDS)
        failures = still_failing

    log.info("%s", "=" * 78)
    log.info("Deleted %d | already done %d | still failing %d", deleted, skipped,
             len(failures))
    for code, sku, detail in failures[:25]:
        log.warning("  UNRESOLVED %s %s - %s", code, sku, detail)
    if len(failures) > 25:
        log.warning("  ... and %d more (see %s)", len(failures) - 25,
                    os.path.basename(LEDGER_FILE))
    log.info("Amazon queues deletions - allow a few minutes, then re-run this "
             "script to verify the marketplace is clean.")
    log.info("%s", "=" * 78)


# =============================================================================
#  MAIN
# =============================================================================

def parse_args(argv: list) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Delete every Amazon listing that is not in the master product list.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--markets", default="",
                   help="comma list of marketplace codes (default: all configured)")
    p.add_argument("--master", default=MASTER_FILE,
                   help="path to the master product xlsx")
    p.add_argument("--live", action="store_true",
                   help="actually delete. Without this nothing is written to Amazon.")
    p.add_argument("--yes", action="store_true",
                   help="skip the typed confirmation on a large live run")
    p.add_argument("--limit", type=int, default=None,
                   help="delete at most N listings per marketplace (canary run)")
    p.add_argument("--refresh", action="store_true",
                   help="re-generate reports instead of reusing today's cache")
    p.add_argument("--no-sheet", action="store_true",
                   help="skip the Google Sheet read; protect parents by KEEP_REGEX only")
    p.add_argument("--report-only", action="store_true",
                   help="write the audit CSVs and stop (no deletion samples)")
    return p.parse_args(argv)


def resolve_markets(spec: str) -> list:
    if spec.strip():
        codes = [c.strip().upper() for c in spec.split(",") if c.strip()]
        unknown = [c for c in codes if c not in MARKETPLACES]
        if unknown:
            raise SystemExit(f"Unknown marketplace code(s): {', '.join(unknown)}. "
                             f"Known: {', '.join(MARKETPLACES)}")
    else:
        codes = list(MARKETPLACES)

    usable = [c for c in codes if REFRESH_TOKENS.get(MARKETPLACES[c]["token"])]
    for c in codes:
        if c not in usable:
            log.warning("%s skipped - no refresh token for auth group %r",
                        c, MARKETPLACES[c]["token"])
    if not usable:
        raise SystemExit("No marketplace has a usable refresh token - check .env.")
    return usable


def confirm(total: int, results: list) -> bool:
    stocked = [d for r in results for d in r["stocked"]]
    unknown = [r["code"] for r in results if not r["fba_known"]]
    print()
    print(f"  About to PERMANENTLY DELETE {total} Amazon listing(s).")
    print("  This removes the offers; the ASINs/detail pages remain.")
    if stocked:
        print(f"  WARNING: {len(stocked)} of them hold {sum(d['fba_qty'] for d in stocked)} "
              f"FBA unit(s), which will be stranded as unfulfillable inventory:")
        for d in sorted(stocked, key=lambda d: -d["fba_qty"])[:10]:
            print(f"      {d['sku']:<28} {d['fba_qty']} unit(s)")
        if len(stocked) > 10:
            print(f"      ... and {len(stocked) - 10} more (see the plan CSV)")
    if unknown:
        print(f"  NOTE: FBA stock could not be read for {', '.join(unknown)} - "
              f"units there may be stranded without warning.")
    print()
    try:
        answer = input('  Type DELETE to proceed (anything else aborts): ').strip()
    except EOFError:
        return False
    return answer == "DELETE"


def run(argv: list) -> int:
    args = parse_args(argv)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = now_stamp()

    log.info("=== Catalogue enforcement: master list is the only source of truth ===")
    log.info("Mode: %s", "LIVE - deletions WILL be sent" if args.live
             else "PREVIEW - nothing will be written to Amazon")

    master = load_master(args.master)
    parents = set() if args.no_sheet else load_sheet_parents(master)
    if args.no_sheet:
        log.warning("--no-sheet: variation parents protected only by %s",
                    KEEP_REGEX.pattern)

    codes = resolve_markets(args.markets)
    log.info("Marketplaces: %s", ", ".join(codes))

    results = []
    for code in codes:
        res = audit_market(code, master, parents, refresh=args.refresh,
                           use_pattern=args.no_sheet)
        if res is not None:
            results.append(res)

    if not results:
        log.error("No marketplace could be audited - nothing to do.")
        return 1

    plan_csv = write_plan_csv(results, stamp)
    coverage_csv = write_coverage_csv(results, master, stamp)
    print_summary(results, master)
    log.info("Deletion plan : %s", plan_csv)
    log.info("Coverage matrix: %s", coverage_csv)

    total_missing = sum(len(r["missing"]) for r in results)
    if total_missing:
        log.warning("%d master SKU/marketplace combinations are MISSING - this is why "
                    "FBA shipment plans report 'MSKU not found'. See the coverage CSV; "
                    "run upload_listings.py to create them.", total_missing)

    total = sum(len(r["delete"]) for r in results)
    if not total:
        log.info("Nothing to delete - every live listing is in the master list.")
        return 0

    if args.report_only:
        log.info("--report-only: stopping after the audit.")
        return 0

    print_samples(results)

    if not args.live:
        log.info("PREVIEW ONLY - re-run with --live to delete these %d listing(s). "
                 "Try `--markets UK --live --limit 5` first.", total)
        return 0

    planned = sum(min(len(r["delete"]), args.limit) if args.limit else len(r["delete"])
                  for r in results)
    if planned > CONFIRM_THRESHOLD and not args.yes and not confirm(planned, results):
        log.info("Aborted - nothing was deleted.")
        return 1

    run_deletions(results, args.limit)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(run(sys.argv[1:]))
    except KeyboardInterrupt:
        log.info("Interrupted - progress is in %s, re-run to resume.",
                 os.path.basename(LEDGER_FILE))
        sys.exit(130)
