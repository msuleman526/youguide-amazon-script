#!/usr/bin/env python3
"""
Market activation — keep ONE marketplace selling, deactivate the rest
=====================================================================

    ####################################################################
    #  BROKEN BY DESIGN — DO NOT RUN AGAINST A REGION YOU SELL IN.     #
    #                                                                  #
    #  This script CANNOT deactivate one marketplace and leave another #
    #  selling. `fulfillment_availability` — the attribute --method    #
    #  fbm0 patches — has NO marketplace_id selector, so it is a       #
    #  SKU-LEVEL value shared across the WHOLE SP-API region. Patching #
    #  it against amazon.ie changes the offer on amazon.co.uk too.     #
    #                                                                  #
    #  Verified live 2026-08-18: reading AALAND-ISLANDS-3GB-15D in     #
    #  three marketplaces returned the identical offer in all three -  #
    #      UK: [('DEFAULT', 0)]  status ['DISCOVERABLE']               #
    #      IE: [('DEFAULT', 0)]  status ['DISCOVERABLE']               #
    #      DE: [('DEFAULT', 0)]  status ['DISCOVERABLE']               #
    #                                                                  #
    #  A live run that day switched off 9 EU marketplaces and took the #
    #  UK down with them: active UK listings 119 -> 23, and 657 UK     #
    #  non-parent listings landed on DEFAULT at quantity 0 - every one #
    #  of them a SKU patched in some OTHER marketplace. The UK was     #
    #  restored with:                                                  #
    #      fbm_to_fba.py --markets UK --excel <list> --live            #
    #  which, being region-wide in the same way, turned the EU offers  #
    #  back on as well.                                                #
    #                                                                  #
    #  NEVER_DEACTIVATE / is_protected() did NOT prevent this and      #
    #  cannot: they gate the marketplace a CALL IS AIMED AT, and the   #
    #  ledger correctly shows zero UK calls. The damage was collateral.#
    #  Any replacement method must be proved by READING THE PROTECTED  #
    #  MARKETPLACE BACK after a canary - the original verification     #
    #  checked amazon.ie alone and so missed this entirely.            #
    ####################################################################

Sells only where you want to sell. Every marketplace is listed in a per-region
array below; only the codes in ACTIVE_MARKETS stay live. Everything else is
DEACTIVATED — the offer is switched off, the listing is NOT deleted, so the SKU,
its ASIN, its content, its reviews and its sales history all survive and can be
switched back on later.

That much still holds: nothing here destroys a listing. What does NOT hold is the
per-marketplace targeting the whole design rests on — see the banner above.

WHAT IT DOES
------------
  1. Pulls the full merchant listings report for every marketplace (the same
     GET_MERCHANT_LISTINGS_ALL_DATA source purge_listings.py / fetch_listings.py
     use, because searchListingsItems will not paginate for this account).
     Reports are cached per marketplace per day — --refresh forces a re-pull.
  2. Reads the product list from the GOOGLE SHEET (read only — no cell is ever
     written) and expands each row by its countries column: one sheet row is a
     SEPARATE PRODUCT in every country it names, so 459 rows x 10 marketplaces
     is 4,590 products, and each one is switched off individually. Anything live
     on Amazon that the sheet does not know about is picked up too.
     Each row carries the eSIM wholesale cost (column G) and therefore the price
     the pricing policy says it SHOULD sell for in that marketplace's currency.
  3. Deactivates EVERY offer outside ACTIVE_MARKETS, whatever status the report
     shows it in (preview by default). An "Inactive" listing is still switched
     off, because inactive usually means out-of-stock, not withdrawn — it comes
     back the moment stock lands. --skip-inactive opts out for faster re-runs.
  4. Writes ONE LOCAL EXCEL FILE with a row per SKU x marketplace:
        sku | name | product_id (ASIN) | eSIM cost | price after formula |
        price on Amazon | price OK? (GREEN/RED) | status (GREEN/RED)
     GREEN price  = Amazon is charging what the formula says (within tolerance).
     GREEN status = the listing is in the state you asked for (UK active,
                    everywhere else deactivated). RED = it is not.

HOW "DEACTIVATE" IS DONE
------------------------
Amazon has no "close listing" call in the Listings Items API. Two reversible
ways to switch an offer off, chosen with --method:

  fbm0   (default)  Delete the FBA entry from `fulfillment_availability`, leaving
                    the merchant channel at quantity 0 — the listing goes
                    "Inactive (Out of Stock)". Reverse it with fbm_to_fba.py.
                    DEACTIVATES THE SKU IN EVERY MARKETPLACE IN THE REGION, not
                    just the one passed in marketplaceIds — the attribute carries
                    no marketplace_id, so there is nothing to target with. This
                    is the failure in the banner at the top of the file, and it
                    is the reason the script cannot do what its name says.
  offer             Delete the `purchasable_offer` attribute. DOES NOT WORK, and
                    the flag is kept only so the finding is not lost: verified
                    live on amazon.de 2026-08-18, three SKUs. Amazon ACCEPTED the
                    patch and the attribute really is gone, but the OFFER keeps
                    its price and the listing stays BUYABLE / DISCOVERABLE
                    (AALAND-ISLANDS-3GB-15D still sold at EUR 8.99 eight minutes
                    later). Deleting the attribute does not withdraw the offer.

Neither call deletes the SKU. deleteListingsItem is never used by this script.

SAFETY
------
  * Previews by default. Nothing is written to Amazon without --live.
  * The Google Sheet is READ ONLY — the product list is read from it, but not one
    cell is ever written back. The output goes to a LOCAL .xlsx.
    (--from-xlsx reads a downloaded copy instead and skips Google entirely.)
  * upload_listings.py is NOT modified. This script imports it for the
    credentials, the marketplace registry and the pricing formula, so the two can
    never disagree about a price, but the sync loop is left exactly as it is.
  * Variation PARENTS are never touched — they carry no offer, and switching one
    off would dissolve the variation family.
  * Marketplaces in ACTIVE_MARKETS are read-only, always — but read the banner:
    no call is ever AIMED at the UK, and the UK still went down on 2026-08-18,
    because the patched attribute is region-wide. This guard is real but it is
    not the protection it was believed to be.
  * Every live change is journalled to logs/deactivate_ledger.csv and skipped on
    re-run, so an interrupted run resumes instead of re-issuing thousands of calls.
  * --limit N turns any run into a canary: N deactivations per marketplace.

USAGE
-----
    python market_activation.py                      # preview + Excel, no writes
    python market_activation.py --markets DE         # one marketplace
    python market_activation.py --markets DE --live --limit 3   # canary
    python market_activation.py --live               # the real run
    python market_activation.py --live --method fbm0 # if `offer` is rejected
    python market_activation.py --report-only        # just the Excel, skip patching
    python market_activation.py --refresh            # ignore today's cached reports
    python market_activation.py --live --skip-inactive   # only touch Active offers
    python market_activation.py --from-xlsx          # offline: read Amazon Listings.xlsx

Credentials, the marketplace registry and the pricing formula all come from
upload_listings, so there is still exactly one place where secrets, IDs and
prices are defined.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sp_api.base import SellingApiException

import upload_listings as ul
from upload_listings import (
    COL_MARKETPLACE,
    COL_MARKET_STATE,
    COL_PARENT_SKU,
    COL_PRICE,
    COL_SKU,
    COL_TITLE,
    DEFAULT_MARKETPLACE_CODE,
    FBM_CHANNEL,
    FX_MARKUP_PCT,
    MARKETPLACES,
    PRICING_CURRENCY,
    PRODUCT_TYPE,
    REFRESH_TOKENS,
    convert_amount,
    fba_channel_for,
    get_client_for_marketplace,
    market_price,
    parse_countries,
    retail_price_raw,
)
from purge_listings import fetch_report, norm_sku, parse_report

# The console on Windows defaults to cp1252 and dies on any non-ASCII log line.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:      # pragma: no cover - very old Python
    pass


# =============================================================================
#  WHICH MARKETPLACES SELL  —  the two lists that decide everything
# =============================================================================

# Every marketplace the account can reach, grouped by SP-API region. These are
# the arrays to edit when a marketplace is added or dropped; the codes must
# exist in upload_listings.MARKETPLACES.
REGION_COUNTRIES = {
    "EU": ["UK", "IE", "FR", "DE", "IT", "ES", "NL", "BE", "SE", "PL"],
    "NA": ["US", "CA", "MX"],
    "FE": ["JP", "AU"],
}

# The ONLY marketplaces allowed to sell. Everything in REGION_COUNTRIES that is
# not in here gets deactivated. Add a code here to turn a market back on (then
# let upload_listings.py re-push it).
ACTIVE_MARKETS = ["UK"]

# Marketplaces this script must NEVER switch off, whatever anyone edits above.
# The client's instruction is "deactivate everywhere EXCEPT the United Kingdom",
# so the UK is protected on its own, independently of ACTIVE_MARKETS: emptying
# or mistyping that list still cannot aim a patch at amazon.co.uk.
#
# READ THIS BEFORE TRUSTING IT. Aiming no patch at the UK is NOT the same as the
# UK being safe. --method fbm0 patches a region-wide attribute, so on 2026-08-18
# the UK was switched off (119 active listings -> 23) without a single UK call
# ever being made — the ledger is still empty of UK rows. See the file banner.
NEVER_DEACTIVATE = {"UK"}


def is_protected(code: str) -> bool:
    """True when `code` must keep selling — checked before every single patch.

    Guards the marketplace a call is AIMED AT. It cannot guard against collateral
    damage from an attribute whose scope is the whole region, which is exactly how
    the UK went down on 2026-08-18. Do not read a passing check here as proof the
    protected marketplace survived — read the marketplace itself back.
    """
    return code in NEVER_DEACTIVATE or code in ACTIVE_MARKETS


# =============================================================================
#  CONFIGURATION
# =============================================================================

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(HERE, "logs")
LEDGER_FILE = os.path.join(OUTPUT_DIR, "deactivate_ledger.csv")

# Offline fallback for --from-xlsx: a downloaded copy of the Google Sheet. Keep
# it fresh (Google Sheets → File → Download → Microsoft Excel) or the eSIM cost
# in column G is stale and "price after formula" prices off old numbers. The
# default run reads the live sheet instead, so this only matters offline.
MASTER_XLSX = os.path.join(HERE, "Amazon Listings.xlsx")

# How far Amazon's live price may drift from the formula price before the cell
# turns RED. Charm rounding lands on .99 in each currency, so an exact match is
# normal; the tolerance only absorbs FX drift between this run and the last push.
PRICE_TOLERANCE_PCT = 1.0
PRICE_TOLERANCE_ABS = 0.02

# Echoed back unchanged so mode=VALIDATION_PREVIEW (used by preview runs) does
# not report them missing — it validates a patch in isolation, without merging it
# against the live listing. See the fbm_to_fba.py docstring.
ECHO_ATTRS = ["batteries_required"]

# The Excel file is rewritten after every marketplace and the ledger flushed
# every LEDGER_FLUSH rows, so a run that is stopped (or watched impatiently)
# still has an up-to-date file on disk instead of nothing until the very end.
LEDGER_FLUSH = 25
PROGRESS_EVERY = 50

SLEEP_BETWEEN_CALLS = 0.6    # Listings Items patch is 5 rps; stay well under it
API_RETRIES = 3
RETRY_BACKOFF = 8.0

# Report statuses that mean "this offer is live and needs switching off".
LIVE_STATUSES = {"active"}
# Statuses that mean it is already off — no call needed.
OFF_STATUSES = {"inactive", "incomplete"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("activation")

# ── Excel fills ──────────────────────────────────────────────────────────────
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="EDEDED")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")


# =============================================================================
#  HELPERS
# =============================================================================

def now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def with_retry(fn, *args, **kwargs):
    """Call fn, retrying transient network failures (flaky TLS / read timeouts)."""
    for attempt in range(1, API_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except SellingApiException:
            raise                       # real API errors are the caller's problem
        except Exception as exc:
            if attempt == API_RETRIES:
                raise
            log.warning("Network hiccup (%s) - retry %d/%d",
                        exc.__class__.__name__, attempt, API_RETRIES)
            time.sleep(RETRY_BACKOFF)


def to_float(value) -> Optional[float]:
    """Parse a report price cell. Returns None for blank/garbage, never raises."""
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_issues(issues: list) -> str:
    """Flatten Amazon's issue list into one readable line."""
    parts = []
    for i in issues or []:
        msg = (i.get("message", "") or "").replace("\n", " ").strip()
        parts.append(f"[{i.get('severity', '?')} {i.get('code', '?')}] {msg}")
    return " ; ".join(parts)


# =============================================================================
#  THE MASTER SHEET — SKU -> name + eSIM cost
# =============================================================================

def parse_market_state(text: str) -> dict:
    """Column W -> {market code: {asin, price, cur, err}}.

    The sync loop packs one entry per marketplace in there, e.g.
        {"BE": {"asin": "B0H6MFJ63H", "cur": "EUR", "price": 31.99, ...},
         "DE": {"asin": "", "err": "...untergeordneten ASIN...", "hash": ""}}
    It is the only place the per-market FAILURE REASON is recorded, so it fills
    in the ASIN and price for markets the listings report has no row for, and
    explains why. Anything unparseable is ignored — older sheets held other
    things in this column.
    """
    text = (text or "").strip()
    if not text.startswith("{"):
        return {}
    try:
        data = json.loads(text)
    except (ValueError, TypeError):
        return {}
    if not isinstance(data, dict):
        return {}
    return {str(k).upper(): v for k, v in data.items() if isinstance(v, dict)}


def load_sheet(from_xlsx: Optional[str] = None) -> tuple[dict, set]:
    """Read the master product list into ({norm_sku: {...}}, {parent norm_skus}).

    Reads the live Google Sheet — READ ONLY, it never writes a cell — because
    that is where the client's product list actually lives. --from-xlsx reads a
    local workbook of the same shape instead, for an offline run.

    Each product carries its COUNTRIES (column V), because one sheet row is one
    product in EVERY country it names: 459 rows x 10 marketplaces is 4,590
    separate Amazon products, and that is what gets deactivated one by one.
    """
    if from_xlsx:
        if not os.path.exists(from_xlsx):
            raise SystemExit(
                f"Master workbook not found: {from_xlsx}\n"
                f"Download the sheet as .xlsx to that path, point --from-xlsx at "
                f"another file, or pass --from-sheet to read the Google Sheet.")
        wb = openpyxl.load_workbook(from_xlsx, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        source = f"{os.path.basename(from_xlsx)} (tab {ws.title!r})"
    else:
        rows = ul.get_worksheet().get_all_values()[1:]   # row 1 is the header
        source = "Google Sheet"

    products: dict = {}
    parents: set = set()
    unknown: dict = {}
    for row in rows:
        def cell(idx):
            return str(row[idx]).strip() if len(row) > idx and row[idx] is not None else ""

        sku = cell(COL_SKU)
        if not sku:
            continue
        parent = cell(COL_PARENT_SKU)
        if parent:
            parents.add(norm_sku(parent))
        # Column V: a single code, a comma list, a region, or ALL. Blank falls
        # back to the home marketplace, exactly as upload_listings does.
        countries, bad = parse_countries(cell(COL_MARKETPLACE))
        if bad:
            unknown.setdefault(", ".join(bad), 0)
            unknown[", ".join(bad)] += 1
        products[norm_sku(sku)] = {
            "sku": sku,
            "title": cell(COL_TITLE),
            "cost": to_float(cell(COL_PRICE)) or 0.0,
            "countries": countries or [DEFAULT_MARKETPLACE_CODE],
            "state": parse_market_state(cell(COL_MARKET_STATE)),
        }

    for tokens, count in unknown.items():
        log.warning("Countries column: %d row(s) name %r, which is not a known "
                    "marketplace - ignored", count, tokens)
    pairs = sum(len(p["countries"]) for p in products.values())
    log.info("Master list: %d SKUs x their countries = %d products, plus %d "
             "variation parents, from %s", len(products), pairs, len(parents), source)
    return products, parents


def price_chain(cost_usd: Optional[float], code: str) -> dict:
    """The whole price story for one product in one country, step by step.

        esim      the eSIM Access wholesale cost, USD (sheet column G)
        formula   after the tier multiplier + minimum-profit safeguard, in
                  PRICING_CURRENCY (GBP) — still unrounded
        converted that GBP figure at today's rate in the local currency, before
                  the .99 is applied (identical to `formula` for the UK)
        final     what the shopper pays: `converted` rounded up to the local .99
                  — exactly what market_price() sends to Amazon

    Every value is None when a rate is missing, so the row says so rather than
    showing a number under the wrong currency.
    """
    currency = MARKETPLACES[code]["currency"]
    out = {"esim": cost_usd, "formula": None, "converted": None,
           "final": None, "currency": currency}
    if not cost_usd or cost_usd <= 0:
        return out
    raw = retail_price_raw(cost_usd)
    if raw is None:
        return out
    out["formula"] = round(raw, 2)
    local = (raw if currency == PRICING_CURRENCY
             else convert_amount(raw, PRICING_CURRENCY, currency))
    if local is not None:
        out["converted"] = round(local * (1 + FX_MARKUP_PCT / 100.0), 2)
    out["final"] = market_price(cost_usd, code)
    return out


# =============================================================================
#  THE LEDGER — so an interrupted live run resumes
# =============================================================================

def load_ledger() -> set:
    """{(market, sku)} already deactivated by a previous live run."""
    if not os.path.exists(LEDGER_FILE):
        return set()
    done = set()
    with open(LEDGER_FILE, encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            if (row.get("result") or "").upper() == "DEACTIVATED":
                done.add((row.get("market", ""), row.get("sku", "")))
    if done:
        log.info("Ledger: %d listings already deactivated in earlier runs - skipped", len(done))
    return done


def append_ledger(entries: list) -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    new_file = not os.path.exists(LEDGER_FILE)
    with open(LEDGER_FILE, "a", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["when", "market", "sku", "asin",
                                                "method", "result", "issues"])
        if new_file:
            writer.writeheader()
        writer.writerows(entries)


# =============================================================================
#  DEACTIVATION
# =============================================================================

def build_patches(payload: dict, method: str, market: str) -> tuple[list, str]:
    """(patches, why-nothing-to-do) for one live listing.

    `offer` deletes the price so the listing cannot be bought; `fbm0` moves the
    offer to the merchant channel with zero stock. Both leave the SKU in place.
    """
    attrs = payload.get("attributes", {}) or {}

    if method == "offer":
        live_offer = attrs.get("purchasable_offer")
        if not live_offer:
            return [], "no purchasable_offer — already has no price"
        # Deleted BY VALUE: purchasable_offer is selector-keyed on
        # marketplace_id/currency, and a valueless delete is rejected on those.
        return [{"op": "delete", "path": "/attributes/purchasable_offer",
                 "value": live_offer}], ""

    # fbm0 — merchant channel, zero quantity.
    fa = attrs.get("fulfillment_availability") or []
    codes = [e.get("fulfillment_channel_code", "") for e in fa]
    fba = fba_channel_for(market)
    patches = [{"op": "add", "path": "/attributes/fulfillment_availability",
                "value": [{"fulfillment_channel_code": FBM_CHANNEL, "quantity": 0}]}]
    # Order matters: the schema wants at least one entry, so add before deleting.
    if fba in codes:
        patches.append({"op": "delete", "path": "/attributes/fulfillment_availability",
                        "value": [{"fulfillment_channel_code": fba}]})
    return patches, ""


def deactivate_one(client, seller_id: str, marketplace_id: str, market: str,
                   sku: str, method: str, live: bool) -> dict:
    """Read a SKU and switch its offer off. Returns {result, issues, asin}."""
    out = {"result": "", "issues": "", "asin": ""}

    # Hard stop. The caller already skips active marketplaces; this is the second
    # lock, so no refactor, flag or typo can ever switch off the market that must
    # keep selling.
    if is_protected(market):
        raise RuntimeError(f"refusing to deactivate {sku} on {market}: "
                           f"{market} is protected (NEVER_DEACTIVATE / ACTIVE_MARKETS)")

    try:
        resp = with_retry(
            client.get_listings_item,
            sellerId=seller_id, sku=sku, marketplaceIds=[marketplace_id],
            includedData=["summaries", "attributes", "offers",
                          "fulfillmentAvailability", "issues"],
        )
    except SellingApiException as exc:
        if getattr(exc, "code", None) == 404:
            out.update(result="NOT_LISTED", issues="SKU not found in this marketplace")
        else:
            out.update(result="READ_ERROR", issues=str(exc))
        return out
    except Exception as exc:
        out.update(result="READ_TIMEOUT", issues=f"{type(exc).__name__}: {exc}")
        return out

    payload = resp.payload or {}
    summaries = payload.get("summaries") or []
    if summaries:
        out["asin"] = summaries[0].get("asin", "")

    # ── Already done? ────────────────────────────────────────────────────────
    # The OFFER view is the authoritative live state, not the attributes list: a
    # converted SKU keeps its old channel in `attributes` forever (the same trap
    # fbm_to_fba.py documents), so checking attributes would re-patch every SKU
    # on every run. If the offer no longer sits on the FBA channel, this listing
    # is already switched off and costs nothing further.
    if method == "fbm0":
        offer_channels = [e.get("fulfillmentChannelCode")
                          for e in (payload.get("fulfillmentAvailability") or [])]
        if fba_channel_for(market) not in offer_channels:
            out.update(result="ALREADY_OFF",
                       issues=f"offer already on {','.join(offer_channels) or 'no channel'}")
            return out

    patches, nothing = build_patches(payload, method, market)
    if nothing:
        out.update(result="ALREADY_OFF", issues=nothing)
        return out

    attrs = payload.get("attributes", {}) or {}
    for name in ECHO_ATTRS:
        if attrs.get(name):
            patches.append({"op": "replace", "path": f"/attributes/{name}",
                            "value": attrs[name]})

    params = {
        "sellerId": seller_id, "sku": sku, "marketplaceIds": [marketplace_id],
        "includedData": ["issues"],
        "body": {"productType": PRODUCT_TYPE, "patches": patches},
    }
    if not live:
        params["mode"] = "VALIDATION_PREVIEW"

    try:
        presp = with_retry(client.patch_listings_item, **params)
    except SellingApiException as exc:
        out.update(result="API_ERROR", issues=str(exc))
        return out
    except Exception as exc:
        # A timeout leaves the state unconfirmed — the next run's report says
        # what actually happened, so never record it as done.
        out.update(result="PATCH_TIMEOUT", issues=f"{type(exc).__name__}: {exc}")
        return out

    data = presp.payload or {}
    status = data.get("status", "UNKNOWN")
    issues = format_issues(data.get("issues", []))
    if status in ("ACCEPTED", "VALID"):
        # ACCEPTED means QUEUED, not applied — the next report pass is the proof.
        out.update(result="DEACTIVATED" if live else "WOULD_DEACTIVATE", issues=issues)
    else:
        out.update(result="FAILED", issues=f"{status} | {issues}")
    return out


# =============================================================================
#  ONE MARKETPLACE
# =============================================================================

def process_market(code: str, products: dict, parents: set, args,
                   ledger: set) -> tuple[list, list]:
    """Audit one marketplace and (unless it is active/report-only) switch it off.

    Returns (excel rows, ledger entries).
    """
    cfg = MARKETPLACES[code]
    currency = cfg["currency"]        # what the shopper pays in — PLN, SEK, ...
    is_active_market = is_protected(code)

    # The report supplies Amazon's side of each row — live price, status, ASIN.
    # It is NOT the work list; the sheet is. A failed report only costs detail.
    text = fetch_report(code, refresh=args.refresh)
    listings = parse_report(code, text) if text else []
    if text is None:
        log.warning("[%s] listings report unavailable - prices and statuses will "
                    "read UNKNOWN, but every sheet product is still processed", code)
    listing_by_key = {i["key"]: i for i in listings}

    # ── The work list ────────────────────────────────────────────────────────
    # One entry per SHEET PRODUCT that names this country in column V — that is
    # what "456 products, uploaded separately in each country" means, and it is
    # what gets switched off one by one. Anything live on Amazon that the sheet
    # does not know about is appended after it, because "deactivate everything
    # except the UK" covers rogue leftovers too (see purge_listings.py).
    targets, seen = [], set()
    for key, product in products.items():
        if code in product["countries"]:
            targets.append((key, product, listing_by_key.get(key)))
            seen.add(key)
    extra = [(k, None, i) for k, i in listing_by_key.items() if k not in seen]
    targets.extend(extra)

    log.info("[%s] %d products from the sheet + %d live on Amazon only = %d rows"
             " | %s", code, len(seen), len(extra), len(targets),
             "ACTIVE MARKET - read only" if is_active_market else "deactivating")

    client = None if (is_active_market or args.report_only) else get_client_for_marketplace(code)
    seller_id = ""
    if client is not None:
        ul.set_active_marketplace(code)
        seller_id = ul.SELLER_ID
        if not seller_id:
            log.error("[%s] no seller id for auth group %r - marketplace skipped",
                      code, cfg["token"])
            client = None

    rows, ledger_entries, done = [], [], 0
    for key, product, item in targets:
        listed = item is not None
        # A sheet product with no listing here still gets a row (and still gets
        # a call — the report can be a day stale and Amazon is the authority).
        if not listed:
            item = {"sku": product["sku"], "key": key, "asin": "", "name": "",
                    "price": "", "status": "", "channel": "", "opened": ""}
        source = ("SHEET + AMAZON" if (product and listed)
                  else "SHEET (not listed here)" if product else "AMAZON ONLY")
        report_status = (item["status"] or "").strip() or ("UNKNOWN" if listed
                                                           else "NOT LISTED")
        status_key = report_status.lower()
        is_parent = key in parents

        # ── What the price SHOULD be here, step by step ──────────────────────
        cost = product["cost"] if product else None
        chain = price_chain(cost, code)
        expected = chain["final"]

        # ── What Amazon actually shows ───────────────────────────────────────
        # The report is Amazon's own truth and wins. Column W is the sync loop's
        # per-market record and fills the gaps — including the reason a market
        # has no listing at all (its "err" text).
        state = (product or {}).get("state", {}).get(code, {})
        amazon_price = to_float(item["price"])
        if amazon_price is None:
            amazon_price = to_float(state.get("price"))
        if not item["asin"]:
            item["asin"] = str(state.get("asin") or "")
        state_error = str(state.get("err") or "").strip()

        if is_parent or amazon_price is None:
            price_check = "N/A"          # parents carry no offer
        elif product is None:
            price_check = "NOT IN SHEET"  # old-naming leftover — see purge_listings.py
        elif expected is None:
            price_check = "NO COST IN SHEET" if not cost else "NO FX RATE"
        else:
            # Named in the currency the shopper actually pays in, so the cell
            # reads "TOO LOW by 15.00 PLN" rather than an abstract mismatch.
            gap = amazon_price - expected
            tol = max(PRICE_TOLERANCE_ABS, expected * PRICE_TOLERANCE_PCT / 100.0)
            if abs(gap) <= tol:
                price_check = "OK"
            else:
                price_check = (f"TOO LOW by {abs(gap):,.2f} {currency}" if gap < 0
                               else f"TOO HIGH by {abs(gap):,.2f} {currency}")

        # ── What should happen to this listing ───────────────────────────────
        # Parents first, and in EVERY marketplace: a variation parent carries no
        # offer, so "not selling" is its correct state, not a fault to flag.
        if is_parent:
            plan = "PARENT - SKIP"
        elif is_active_market:
            plan = "KEEP ACTIVE"
        elif not listed and not args.verify_unlisted:
            # The report says this product was never listed in this marketplace.
            # --verify-unlisted asks Amazon anyway instead of trusting that.
            plan = "NOT LISTED HERE"
        elif args.skip_inactive and status_key in OFF_STATUSES:
            plan = "ALREADY OFF"
        else:
            # Every product is attempted, whatever the report says its status is:
            # the report is up to a day old, "Inactive" can mean out-of-stock
            # rather than switched off, and a listing that still carries a price
            # comes straight back the moment stock arrives. The live read inside
            # deactivate_one() is what decides there is nothing left to do.
            plan = "DEACTIVATE"

        result, issues = "", ""
        if plan == "KEEP ACTIVE":
            result = "ACTIVE" if status_key in LIVE_STATUSES else f"NOT SELLING ({report_status})"
        elif plan == "PARENT - SKIP":
            result = "PARENT"
        elif plan == "ALREADY OFF":
            result = "ALREADY_OFF"
        elif plan == "NOT LISTED HERE":
            result = "NOT_LISTED"
        elif args.report_only or client is None:
            result = "NOT ATTEMPTED"
        elif (code, item["sku"]) in ledger:
            result = "DEACTIVATED"
            issues = "already done in an earlier run (ledger)"
        elif args.limit and done >= args.limit:
            result = "SKIPPED (--limit)"
        else:
            outcome = deactivate_one(client, seller_id, cfg["id"], code,
                                     item["sku"], args.method, args.live)
            result, issues = outcome["result"], outcome["issues"]
            if outcome["asin"]:
                item["asin"] = outcome["asin"]
            done += 1
            level = log.info if result in ("DEACTIVATED", "WOULD_DEACTIVATE",
                                           "ALREADY_OFF") else log.warning
            level("  %-32s %s  %s %s", item["sku"], code, result,
                  f"| {issues[:120]}" if issues else "")
            if args.live:
                ledger_entries.append({
                    "when": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    "market": code, "sku": item["sku"], "asin": item["asin"],
                    "method": args.method, "result": result, "issues": issues[:400],
                })
                if len(ledger_entries) >= LEDGER_FLUSH:
                    append_ledger(ledger_entries)   # survive a Ctrl-C mid-market
                    ledger_entries = []
            if done % PROGRESS_EVERY == 0:
                log.info("[%s] %d/%d done", code, done,
                         sum(1 for t in targets if t is not None))
            time.sleep(SLEEP_BETWEEN_CALLS)

        rows.append({
            "market": code,
            "currency": currency,
            "sku": item["sku"],
            "name": (product["title"] if product else item["name"]) or item["name"],
            "product_id": item["asin"],
            "esim_cost_usd": chain["esim"],
            "formula_price": chain["formula"],
            "converted_price": chain["converted"],
            "final_price": chain["final"],
            "price_on_amazon": amazon_price,
            "price_check": price_check,
            "amazon_status": report_status,
            "plan": plan,
            "result": result,
            "source": source,
            "notes": issues or state_error,
        })

    return rows, ledger_entries


# =============================================================================
#  THE EXCEL FILE
# =============================================================================

COLUMNS = [
    ("market", "Country", 9),
    ("sku", "SKU", 34),
    ("name", "Name", 50),
    ("product_id", "Product ID (ASIN)", 18),
    # The price story, one column per step: cost -> formula -> FX -> shelf price.
    ("esim_cost_usd", "eSIM Access price (USD)", 15),
    ("formula_price", "Formula price (GBP)", 15),
    ("converted_price", "Converted price", 14),
    ("final_price", "Final price", 12),
    ("currency", "Currency", 9),
    ("price_on_amazon", "Price on Amazon", 14),
    ("price_check", "Price OK?", 14),
    ("result", "Status", 20),
    ("amazon_status", "Amazon status", 13),
    ("source", "Source", 21),
    ("notes", "Notes", 60),
]

def price_is_bad(check: str) -> bool:
    """True for a priced row whose price is wrong — the RED price cells."""
    return str(check).startswith("TOO ")


# A result is GREEN when the listing is in the state ACTIVE_MARKETS asks for.
GOOD_RESULTS = {"ACTIVE", "DEACTIVATED", "ALREADY_OFF", "NOT_LISTED"}
PENDING_RESULTS = {"WOULD_DEACTIVATE", "NOT ATTEMPTED", "SKIPPED (--limit)", "PARENT"}


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def write_excel(rows: list, path: str, args) -> None:
    """One row per SKU x marketplace, with the two GREEN/RED verdict columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "listings"

    ws.append([label for _, label, _ in COLUMNS])
    col_of = {key: i + 1 for i, (key, _, _) in enumerate(COLUMNS)}
    for key, _, width in COLUMNS:
        ws.column_dimensions[get_column_letter(col_of[key])].width = width

    check_col = col_of["price_check"]
    status_col = col_of["result"]
    price_cols = [col_of[k] for k in ("esim_cost_usd", "formula_price",
                                      "converted_price", "final_price",
                                      "price_on_amazon")]

    for row in rows:
        ws.append([row.get(key) for key, _, _ in COLUMNS])
        r = ws.max_row

        # Price verdict — GREEN when Amazon charges what the formula says.
        check = row["price_check"]
        cell = ws.cell(row=r, column=check_col)
        cell.fill = (GREEN if check == "OK" else RED if price_is_bad(check)
                     else GREY)   # grey = nothing to compare (parent, unlisted)

        # Listing verdict — GREEN when it is in the state we asked for.
        result = row["result"]
        cell = ws.cell(row=r, column=status_col)
        cell.fill = (GREEN if result in GOOD_RESULTS
                     else AMBER if result in PENDING_RESULTS else RED)

        for col in price_cols:
            ws.cell(row=r, column=col).number_format = "0.00"

    style_header(ws, len(COLUMNS))

    # ── Summary tab ──────────────────────────────────────────────────────────
    sm = wb.create_sheet("summary")
    sm.append(["Market", "Role", "Listings", "Amazon Active", "Deactivated / off",
               "Failed", "Price OK", "Price mismatch"])
    for code in sorted({r["market"] for r in rows}):
        mine = [r for r in rows if r["market"] == code]
        sm.append([
            code,
            "ACTIVE" if is_protected(code) else "deactivate",
            len(mine),
            sum(1 for r in mine if r["amazon_status"].lower() in LIVE_STATUSES),
            sum(1 for r in mine if r["result"] in ("DEACTIVATED", "ALREADY_OFF")),
            sum(1 for r in mine if r["result"] in ("FAILED", "API_ERROR",
                                                   "READ_ERROR", "PATCH_TIMEOUT",
                                                   "READ_TIMEOUT")),
            sum(1 for r in mine if r["price_check"] == "OK"),
            sum(1 for r in mine if price_is_bad(r["price_check"])),
        ])
    for col, width in zip("ABCDEFGH", (10, 12, 10, 15, 18, 9, 10, 16)):
        sm.column_dimensions[col].width = width
    style_header(sm, 8)

    sm.append([])
    sm.append(["Mode", "LIVE" if args.live else "PREVIEW (no writes)"])
    sm.append(["Method", args.method])
    sm.append(["Active markets", ", ".join(ACTIVE_MARKETS)])
    sm.append(["Generated (UTC)", datetime.now(timezone.utc)
               .strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(path)
    log.info("Excel written: %s", path)


# =============================================================================
#  MAIN
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--live", action="store_true",
                   help="actually deactivate. Without this nothing is written.")
    p.add_argument("--method", choices=("fbm0", "offer"), default="fbm0",
                   help="fbm0 (default) = merchant channel with 0 stock, the one that "
                        "actually deactivates; offer = delete purchasable_offer, which "
                        "live testing showed leaves the listing BUYABLE - do not use")
    p.add_argument("--markets", default="",
                   help="comma list of marketplace codes (default: all in REGION_COUNTRIES)")
    p.add_argument("--limit", type=int, default=0,
                   help="stop after N deactivations per marketplace (canary runs)")
    p.add_argument("--report-only", action="store_true",
                   help="build the Excel file only, touch nothing on Amazon")
    p.add_argument("--refresh", action="store_true",
                   help="re-pull the listings reports instead of reusing today's cache")
    p.add_argument("--skip-inactive", action="store_true",
                   help="do not call Amazon for listings the report already shows as "
                        "Inactive/Incomplete (faster re-runs; default is to switch "
                        "off EVERY product regardless of its reported status)")
    p.add_argument("--verify-unlisted", action="store_true",
                   help="ask Amazon about sheet products the report shows as never "
                        "listed in that marketplace, instead of trusting the report")
    p.add_argument("--from-xlsx", nargs="?", const=MASTER_XLSX, default="",
                   help=f"read the product list from a local workbook instead of the "
                        f"Google Sheet (bare flag uses {os.path.basename(MASTER_XLSX)})")
    p.add_argument("--out", default="",
                   help="output .xlsx path (default: market_activation_<stamp>.xlsx)")
    return p.parse_args()


def resolve_markets(requested: str) -> list:
    """Marketplace codes to process: configured, known, and with a token."""
    known = [c for group in REGION_COUNTRIES.values() for c in group]
    unknown = [c for c in known if c not in MARKETPLACES]
    if unknown:
        log.warning("REGION_COUNTRIES names %s, which upload_listings does not know - ignored",
                    ", ".join(unknown))
    codes = [c for c in known if c in MARKETPLACES]

    if requested:
        wanted = [c.strip().upper() for c in requested.replace(";", ",").split(",") if c.strip()]
        bad = [c for c in wanted if c not in codes]
        if bad:
            raise SystemExit(f"Unknown marketplace code(s): {', '.join(bad)}\n"
                             f"Known: {', '.join(codes)}")
        codes = [c for c in codes if c in wanted]

    live_codes, skipped = [], []
    for c in codes:
        (live_codes if REFRESH_TOKENS.get(MARKETPLACES[c]["token"]) else skipped).append(c)
    if skipped:
        log.warning("No refresh token for %s - those marketplaces cannot be read or "
                    "changed and are left out entirely", ", ".join(skipped))
    return live_codes


def run() -> None:
    args = parse_args()

    missing = [c for c in ACTIVE_MARKETS if c not in MARKETPLACES]
    if missing:
        raise SystemExit(f"ACTIVE_MARKETS names unknown marketplace(s): {', '.join(missing)}")

    codes = resolve_markets(args.markets)
    if not codes:
        raise SystemExit("No marketplaces to process.")

    log.info("=== Market activation ===")
    log.info("Mode        : %s", "LIVE - offers WILL be switched off" if args.live
             else "PREVIEW - nothing is written")
    log.info("Method      : %s", args.method)
    log.info("Active      : %s  (kept selling, never touched)", ", ".join(ACTIVE_MARKETS))
    log.info("Protected   : %s  (cannot be deactivated by this script at all)",
             ", ".join(sorted(NEVER_DEACTIVATE)))
    log.info("Deactivating: %s", ", ".join(c for c in codes if not is_protected(c)) or "(none)")

    products, parents = load_sheet(args.from_xlsx or None)
    ledger = load_ledger() if args.live else set()

    out = args.out or os.path.join(HERE, f"market_activation_{now_stamp()}.xlsx")
    all_rows: list = []
    for code in codes:
        rows, entries = process_market(code, products, parents, args, ledger)
        all_rows.extend(rows)
        if entries:
            append_ledger(entries)      # the tail below the last flush
        if all_rows:
            # Rewritten after EVERY marketplace, not once at the end: a two-hour
            # run must not leave you staring at an empty folder, and a run that
            # is stopped early still leaves a complete file for what it covered.
            write_excel(all_rows, out, args)
            log.info("[%s] done - Excel updated: %s (%d rows so far)",
                     code, os.path.basename(out), len(all_rows))

    if not all_rows:
        log.warning("No listings found anywhere - nothing to report.")
        return

    # ── Console summary ──────────────────────────────────────────────────────
    log.info("---------------------------------------------------------------")
    log.info("%-8s %-10s %7s %7s %8s %8s %9s", "MARKET", "ROLE", "TOTAL",
             "ACTIVE", "OFF", "FAILED", "PRICE-BAD")
    for code in codes:
        mine = [r for r in all_rows if r["market"] == code]
        if not mine:
            continue
        log.info("%-8s %-10s %7d %7d %8d %8d %9d", code,
                 "ACTIVE" if is_protected(code) else "deactivate", len(mine),
                 sum(1 for r in mine if r["amazon_status"].lower() in LIVE_STATUSES),
                 sum(1 for r in mine if r["result"] in ("DEACTIVATED", "ALREADY_OFF")),
                 sum(1 for r in mine if r["result"] in ("FAILED", "API_ERROR",
                                                        "READ_ERROR", "PATCH_TIMEOUT",
                                                        "READ_TIMEOUT")),
                 sum(1 for r in mine if price_is_bad(r["price_check"])))
    log.info("---------------------------------------------------------------")

    if not args.live and not args.report_only:
        log.info("PREVIEW only. Re-run with --live to switch the offers off "
                 "(try --markets DE --live --limit 3 first).")
    if args.live:
        log.info("Amazon queues a patch — statuses in the report catch up on the "
                 "next pass. Re-run with --refresh --report-only to confirm.")
    log.info("REMEMBER: upload_listings.py re-pushes every marketplace named in "
             "the sheet's countries column (V). Restrict column V to %s, or the "
             "sync loop will turn these listings back on.", ", ".join(ACTIVE_MARKETS))


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        log.warning("Interrupted - the ledger keeps what was already done.")
    except SellingApiException as exc:
        log.error("SP-API error: %s", exc)
