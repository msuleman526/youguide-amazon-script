#!/usr/bin/env python3
"""
Stock status — what is ACTIVE, what is INACTIVE, and what is OUT OF STOCK
========================================================================
Read-only. Pulls a fresh merchant-listings report AND a fresh FBA inventory
report for every marketplace that has credentials, then writes one workbook:

    Sheets/stock_status_<stamp>.xlsx
        summary   one row per country: active / out of stock / deactivated
        listings  one row per SKU x country, colour-coded by state

WHY TWO REPORTS. The merchant-listings report's `quantity` column is
MERCHANT-fulfilled stock, so it reads 0 for every FBA listing. The real number
lives only in the FBA inventory report (`afn-fulfillable-quantity`), which is
what separates "inactive because it is out of stock" from "inactive because
somebody switched it off" — the distinction that matters after 2026-08-18.

STATES
    ACTIVE         selling now
    OUT OF STOCK   not selling, FBA fulfillable quantity is 0 (the usual case)
    DEACTIVATED    not selling, offer sits on the merchant channel (switched off)
    INACTIVE       not selling for some other reason - read the Amazon status
    PARENT         variation parent; carries no offer, so "not selling" is correct

Only the EU authorization group has a refresh token in .env, so US, CA, MX, JP
and AU cannot be read at all and are reported as skipped, never as "clean".

    python stock_status.py              # all marketplaces with credentials
    python stock_status.py --markets UK # just one
    python stock_status.py --no-refresh # reuse today's cached reports (fast)
"""

from __future__ import annotations

import argparse
import collections
import os
import sys
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import purge_listings as pl
import market_activation as ma
from upload_listings import MARKETPLACES, REFRESH_TOKENS

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="EDEDED")
HEAD = PatternFill("solid", fgColor="1F4E79")
FILLS = {"ACTIVE": GREEN, "OUT OF STOCK": AMBER, "DEACTIVATED": RED,
         "INACTIVE": RED, "PARENT": GREY}

COLUMNS = [("Country", 9), ("SKU", 34), ("Name", 46), ("ASIN", 14),
           ("State", 15), ("Amazon status", 13), ("Channel", 12),
           ("FBA stock", 10), ("Price", 10), ("Why", 42)]


def classify(row: dict, stock: dict) -> tuple[str, str]:
    """One listing -> (state, why). See the module docstring for the states."""
    status = (row["status"] or "").lower()
    if status == "incomplete":
        return "PARENT", "variation parent - carries no offer"
    if status == "active":
        return "ACTIVE", "selling"
    qty = stock.get(row["key"])
    if qty is not None and qty <= 0:
        return "OUT OF STOCK", "FBA fulfillable quantity is 0"
    if row["channel"] == "DEFAULT":
        return "DEACTIVATED", "offer is on the merchant channel, not FBA"
    return "INACTIVE", f"Amazon reports {row['status'] or 'no status'}"


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--markets", default="", help="comma list (default: all with credentials)")
    ap.add_argument("--no-refresh", action="store_true",
                    help="reuse today's cached reports instead of pulling fresh ones")
    ap.add_argument("--out", default="", help="output .xlsx path")
    args = ap.parse_args()

    known = [c for g in ma.REGION_COUNTRIES.values() for c in g]
    if args.markets:
        wanted = {c.strip().upper() for c in args.markets.split(",") if c.strip()}
        known = [c for c in known if c in wanted]

    codes, skipped = [], []
    for c in known:
        (codes if REFRESH_TOKENS.get(MARKETPLACES[c]["token"]) else skipped).append(c)
    if skipped:
        print(f"SKIPPED (no refresh token for their auth group): {', '.join(skipped)}")

    refresh = not args.no_refresh
    rows_out, summary = [], []
    for code in codes:
        text = pl.fetch_report(code, refresh=refresh)
        if not text:
            print(f"[{code}] listings report unavailable - skipped")
            continue
        listings = pl.parse_report(code, text)
        stock = pl.fba_quantities(code, refresh=refresh) or {}

        counts = collections.Counter()
        for r in listings:
            state, why = classify(r, stock)
            counts[state] += 1
            rows_out.append([code, r["sku"], (r["name"] or "")[:90], r["asin"],
                             state, r["status"], r["channel"],
                             stock.get(r["key"]), r["price"], why])
        summary.append([code,
                        "MUST SELL" if code in ma.ACTIVE_MARKETS else "should be off",
                        len(listings), counts["ACTIVE"], counts["OUT OF STOCK"],
                        counts["DEACTIVATED"], counts["INACTIVE"], counts["PARENT"]])
        print(f"[{code}] active={counts['ACTIVE']} out_of_stock={counts['OUT OF STOCK']} "
              f"deactivated={counts['DEACTIVATED']} inactive={counts['INACTIVE']} "
              f"parents={counts['PARENT']}")

    if not rows_out:
        print("Nothing to report.")
        return

    wb = openpyxl.Workbook()
    sm = wb.active
    sm.title = "summary"
    sm.append(["Country", "Role", "Total", "ACTIVE (selling)", "OUT OF STOCK",
               "DEACTIVATED", "INACTIVE (other)", "Parents"])
    for row in summary:
        sm.append(row)
    for i, w in enumerate((10, 16, 8, 16, 14, 14, 16, 9), 1):
        sm.column_dimensions[get_column_letter(i)].width = w
    style_header(sm, 8)
    if skipped:
        sm.append([])
        sm.append(["NOT READ - no credentials", ", ".join(skipped)])

    ws = wb.create_sheet("listings")
    ws.append([c for c, _ in COLUMNS])
    for i, (_, w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in rows_out:
        ws.append(row)
        ws.cell(row=ws.max_row, column=5).fill = FILLS.get(row[4], GREY)
    style_header(ws, len(COLUMNS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    out = args.out or os.path.join(HERE, "Sheets", f"stock_status_{stamp}.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nWritten: {out}  ({len(rows_out)} rows)")


if __name__ == "__main__":
    main()
